"""
MASTER sheet extractor for .xlsm corporate rating files.

Sheet structure:
  - Rows 1–N: key-value pairs where col B = label, col C+ = values
  - Multi-value rows: Industry risk / score / weight each have 1-2 entries
  - "[Scope Credit Metrics]" row marks start of time-series section;
    col C+ of that row holds the years, subsequent rows hold metric values.

Enhancements:
  - Cell-level provenance: every field traces to (sheet, row, col)
  - Schema drift detection: unknown labels flagged as quality issues
  - Business-hash: SHA-256(company_name|business_year_end) for idempotency
  - Content fingerprint: SHA-256 of rating fields for delta detection
"""
from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl


METRIC_LABEL_MAP = {
    "Scope-adjusted EBITDA interest cover": "ebitda_interest_cover",
    "Scope-adjusted debt/EBITDA": "debt_ebitda",
    "Scope-adjusted FFO/debt": "ffo_debt",
    "Scope-adjusted loan/value": "loan_value",
    "Scope-adjusted FOCF/debt": "focf_debt",
    "Liquidity": "liquidity",
}

# Canonical set of all expected MASTER sheet labels — used for drift detection
KNOWN_LABELS: frozenset[str] = frozenset({
    "Rated entity", "CorporateSector", "Rating methodologies applied",
    "Industry risk", "Industry risk score", "Industry weight",
    "Segmentation criteria", "Reporting Currency/Units", "Country of origin",
    "Accounting principles", "End of business year",
    "Business risk profile", "(Blended) Industry risk profile",
    "Competitive Positioning", "Market share", "Diversification",
    "Operating profitability", "Sector/company-specific factors (1)",
    "Sector/company-specific factors (2)", "Financial risk profile",
    "Leverage", "Interest cover", "Cash flow cover", "Liquidity",
    "[Scope Credit Metrics]",
    *METRIC_LABEL_MAP.keys(),
})

VALID_RATING_NOTATIONS = {
    "AAA", "AA+", "AA", "AA-", "A+", "A", "A-",
    "BBB+", "BBB", "BBB-", "BB+", "BB", "BB-",
    "B+", "B", "B-", "CCC+", "CCC", "CCC-", "CC", "C", "D", "SD",
}


@dataclass
class IndustrySegment:
    risk: str
    score: str
    weight: float


@dataclass
class CreditMetrics:
    year: int
    ebitda_interest_cover: float | None = None
    debt_ebitda: float | None = None
    ffo_debt: float | None = None
    loan_value: float | None = None
    focf_debt: float | None = None
    liquidity: float | None = None
    is_estimate: bool = False
    is_stale: bool = False


@dataclass
class ProvenanceRecord:
    field_name: str
    source_sheet: str
    source_row: int
    source_col: str
    raw_value: str
    extracted_value: str


@dataclass
class MasterData:
    filename: str
    file_hash: str
    raw_bytes: bytes

    rated_entity: str = ""
    sector: str = ""
    currency: str = ""
    country: str = ""
    accounting_principles: str = ""
    business_year_end: str = ""
    segmentation_criteria: str = ""

    methodologies: list[str] = field(default_factory=list)
    industry_segments: list[IndustrySegment] = field(default_factory=list)

    business_risk_profile: str = ""
    blended_industry_risk: str = ""
    competitive_positioning: str = ""
    market_share: str = ""
    diversification: str = ""
    operating_profitability: str = ""
    sector_specific_factor_1: str = ""
    sector_specific_factor_2: str = ""
    financial_risk_profile: str = ""
    leverage: str = ""
    interest_cover: str = ""
    cash_flow_cover: str = ""
    liquidity_adjustment: str = ""

    credit_metrics: list[CreditMetrics] = field(default_factory=list)
    quality_issues: list[dict] = field(default_factory=list)
    provenance: list[ProvenanceRecord] = field(default_factory=list)

    # Schema drift
    labels_seen: list[str] = field(default_factory=list)
    unknown_labels: list[str] = field(default_factory=list)

    # Computed after extraction
    business_key: str = ""
    content_fingerprint: str = ""


def _hash(value: str) -> str:
    return hashlib.sha256(value.encode()).hexdigest()


def _hash_file(raw: bytes) -> str:
    return hashlib.sha256(raw).hexdigest()


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str) and value.strip().lower() not in ("no data", "n/a", ""):
        try:
            return float(value.strip())
        except ValueError:
            pass
    return None


def _col_letter(col_index: int) -> str:
    """Convert 1-based column index to Excel column letter."""
    result = ""
    while col_index > 0:
        col_index, remainder = divmod(col_index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _content_fingerprint(data: MasterData) -> str:
    """SHA-256 of all rating/risk fields — used for delta detection."""
    payload = json.dumps({
        "business_risk_profile": data.business_risk_profile,
        "blended_industry_risk": data.blended_industry_risk,
        "competitive_positioning": data.competitive_positioning,
        "market_share": data.market_share,
        "diversification": data.diversification,
        "operating_profitability": data.operating_profitability,
        "sector_specific_factor_1": data.sector_specific_factor_1,
        "sector_specific_factor_2": data.sector_specific_factor_2,
        "financial_risk_profile": data.financial_risk_profile,
        "leverage": data.leverage,
        "interest_cover": data.interest_cover,
        "cash_flow_cover": data.cash_flow_cover,
        "liquidity_adjustment": data.liquidity_adjustment,
        "industry_risks": [
            {"risk": s.risk, "score": s.score, "weight": s.weight}
            for s in data.industry_segments
        ],
    }, sort_keys=True)
    return hashlib.sha256(payload.encode()).hexdigest()


def extract(path: Path) -> MasterData:
    raw = path.read_bytes()
    wb = openpyxl.load_workbook(path, keep_vba=True, data_only=True)

    if "MASTER" not in wb.sheetnames:
        raise ValueError(f"No MASTER sheet in {path.name}")

    ws = wb["MASTER"]
    data = MasterData(filename=path.name, file_hash=_hash_file(raw), raw_bytes=raw)

    industry_risks: list[str] = []
    industry_scores: list[str] = []
    industry_weights: list[float] = []

    in_metrics = False
    years: list[int] = []

    # Iterate with cell objects (not values_only) to capture row/col addresses
    for row_cells in ws.iter_rows():
        if len(row_cells) < 3:
            continue

        label_cell = row_cells[1]  # col B
        if label_cell.value is None:
            continue

        label = str(label_cell.value).strip()
        row_num = label_cell.row

        # Track all labels for drift detection
        if label not in data.labels_seen:
            data.labels_seen.append(label)
        if label not in KNOWN_LABELS and label not in data.unknown_labels:
            data.unknown_labels.append(label)
            data.quality_issues.append({
                "field_name": label,
                "issue_type": "schema_drift",
                "issue_detail": f"Unrecognized label '{label}' in MASTER sheet",
                "severity": "warning",
                "source_sheet": "MASTER",
                "source_row": row_num,
                "source_col": "B",
            })

        values = [c.value for c in row_cells[2:] if c.value is not None]

        def record_provenance(field_name: str, col_idx: int, raw_val: Any, extracted: Any):
            data.provenance.append(ProvenanceRecord(
                field_name=field_name,
                source_sheet="MASTER",
                source_row=row_num,
                source_col=_col_letter(col_idx),
                raw_value=str(raw_val) if raw_val is not None else "",
                extracted_value=str(extracted) if extracted is not None else "",
            ))

        # ── Time-series section ───────────────────────────────────────────
        if label == "[Scope Credit Metrics]":
            in_metrics = True
            years = [int(v) for v in values if isinstance(v, (int, float)) and int(v) == v]
            continue

        if in_metrics:
            attr = METRIC_LABEL_MAP.get(label)
            if attr and years:
                raw_vals = list(row_cells[2: 2 + len(years)])
                for i, (year, cell) in enumerate(zip(years, raw_vals)):
                    raw_val = cell.value
                    fval = _to_float(raw_val)
                    if fval is None and raw_val is not None:
                        data.quality_issues.append({
                            "field_name": f"credit_metrics.{attr}.{year}",
                            "issue_type": "invalid_type",
                            "issue_detail": f"Non-numeric value '{raw_val}' treated as NULL",
                            "severity": "warning",
                            "source_sheet": "MASTER",
                            "source_row": row_num,
                            "source_col": _col_letter(cell.column),
                        })
                    metric = _get_or_create_metric(data.credit_metrics, year)
                    setattr(metric, attr, fval)
                    record_provenance(f"credit_metrics.{attr}.{year}", cell.column, raw_val, fval)
            continue

        # ── Key-value section ─────────────────────────────────────────────
        first_cell = row_cells[2] if len(row_cells) > 2 else None
        first_val = first_cell.value if first_cell else None
        first_col = first_cell.column if first_cell else 3

        def set_field(field_name: str, value: Any):
            extracted = str(value) if value is not None else ""
            record_provenance(field_name, first_col, first_val, extracted)
            return extracted

        match label:
            case "Rated entity":
                data.rated_entity = set_field("rated_entity", first_val)
            case "CorporateSector":
                data.sector = set_field("sector", first_val)
            case "Rating methodologies applied":
                data.methodologies = [str(v) for v in values]
                for i, cell in enumerate(row_cells[2: 2 + len(values)]):
                    record_provenance(f"methodologies[{i}]", cell.column, cell.value, str(cell.value))
            case "Industry risk":
                industry_risks = [str(v) for v in values]
                for i, cell in enumerate(row_cells[2: 2 + len(values)]):
                    record_provenance(f"industry_risk[{i}]", cell.column, cell.value, str(cell.value))
            case "Industry risk score":
                industry_scores = [str(v) for v in values]
            case "Industry weight":
                industry_weights = [float(v) for v in values if isinstance(v, (int, float))]
            case "Segmentation criteria":
                data.segmentation_criteria = set_field("segmentation_criteria", first_val)
            case "Reporting Currency/Units":
                data.currency = set_field("currency", first_val)
            case "Country of origin":
                data.country = set_field("country", first_val)
            case "Accounting principles":
                data.accounting_principles = set_field("accounting_principles", first_val)
            case "End of business year":
                data.business_year_end = set_field("business_year_end", first_val)
            case "Business risk profile":
                data.business_risk_profile = set_field("business_risk_profile", first_val)
            case "(Blended) Industry risk profile":
                data.blended_industry_risk = set_field("blended_industry_risk", first_val)
            case "Competitive Positioning":
                data.competitive_positioning = set_field("competitive_positioning", first_val)
            case "Market share":
                data.market_share = set_field("market_share", first_val)
            case "Diversification":
                data.diversification = set_field("diversification", first_val)
            case "Operating profitability":
                data.operating_profitability = set_field("operating_profitability", first_val)
            case "Sector/company-specific factors (1)":
                data.sector_specific_factor_1 = set_field("sector_specific_factor_1", first_val)
            case "Sector/company-specific factors (2)":
                data.sector_specific_factor_2 = set_field("sector_specific_factor_2", first_val)
            case "Financial risk profile":
                data.financial_risk_profile = set_field("financial_risk_profile", first_val)
            case "Leverage":
                data.leverage = set_field("leverage", first_val)
            case "Interest cover":
                data.interest_cover = set_field("interest_cover", first_val)
            case "Cash flow cover":
                data.cash_flow_cover = set_field("cash_flow_cover", first_val)
            case "Liquidity" if not in_metrics:
                data.liquidity_adjustment = set_field("liquidity_adjustment", first_val)

    # ── Assemble industry segments ────────────────────────────────────────
    for risk, score, weight in zip(industry_risks, industry_scores, industry_weights):
        data.industry_segments.append(IndustrySegment(risk=risk, score=score, weight=weight))

    # ── Computed hashes ───────────────────────────────────────────────────
    data.business_key = _hash(f"{data.rated_entity}|{data.business_year_end}")
    data.content_fingerprint = _content_fingerprint(data)

    return data


def _get_or_create_metric(metrics: list[CreditMetrics], year: int) -> CreditMetrics:
    for m in metrics:
        if m.year == year:
            return m
    m = CreditMetrics(year=year)
    metrics.append(m)
    return m
